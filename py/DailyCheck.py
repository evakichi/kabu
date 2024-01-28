import CommonPackage
import Period
import Token
import Brand
import DailyJpxData
import WeeklyJpxData
import MonthlyJpxData
import TradingCalender
import JpxBrandData
import os
import sys
import math
import json
import requests
import glob
import datetime
import openpyxl

from multiprocessing import Process

refreshToken, idToken = Token.getTokens()

def dailyCheck(idToken,brandData,distance=0):

    headers = {'Authorization': f'Bearer {idToken}'}
    brandCode = brandData.getCode()
    currentPath = CommonPackage.createDir(os.path.join(CommonPackage.dataDir,brandData.getCode()))

    if distance == -1:
        dailyQuotesGet = requests.get(f"https://api.jquants.com/v1/prices/daily_quotes?code={brandCode}", headers=headers)
    else:
        date = Period.Period.getDate(distance).strftime('%Y-%m-%d')
        dailyQuotesGet = requests.get(f"https://api.jquants.com/v1/prices/daily_quotes?code={brandCode}&date={date}", headers=headers)

    dailyQuotesJson = dailyQuotesGet.json()

    if 'daily_quotes' in dailyQuotesJson and len(dailyQuotesJson['daily_quotes'])!=0:
        for dailyQuote in dailyQuotesJson['daily_quotes']:
            current = dailyQuote['Date']
            currentFilePath = os.path.join(currentPath,current+".json")
            if not os.path.exists(currentFilePath):
                with open(currentFilePath, 'w') as f:
                    json.dump(dailyQuote, f)

def readDailyAllJpxData(brandData):
    allDailyJpxDataList = list()
    for i,bd in enumerate(brandData):
        currentPath = os.path.join(CommonPackage.dataDir,bd.getCode())
        fileList = sorted(glob.glob(currentPath+'/*.json',recursive=False))
        dailyJpxDataList = list()
        prev = None
        for j,fl in enumerate(fileList,2):
            with open(fl) as f:
                dailyJpxData = DailyJpxData.DailyJpxData(json.load(f),prev)
            if dailyJpxData.quotes != None:
                dailyJpxDataList.append(dailyJpxData)
                prev = dailyJpxData
        allDailyJpxDataList.append(JpxBrandData.JpxBrandData(bd,dailyJpxDataList))
    return allDailyJpxDataList

def calcAllWeeklyJpxData(allDailyJpxDataList):
    allWeeklyJpxDataList = list()
    for i,adjdl in enumerate(allDailyJpxDataList):
        currentYearWeek = '-1'
        weeklyJpxDataList = list()
        weeklyJpxData = None
        prev = None
        for j,jdl in enumerate(adjdl.getJpxDataList()):
            yearWeek = CommonPackage.getYearWeekNumber(jdl.date)
            if currentYearWeek == yearWeek:
                weeklyJpxData.append(jdl)
            if currentYearWeek != yearWeek:
                if weeklyJpxData != None:
                    weeklyJpxData.reCalc(prev)
                    prev = weeklyJpxData
                    weeklyJpxDataList.append(weeklyJpxData)
                weeklyJpxData = MonthlyJpxData.MonthlyJpxData(yearWeek,jdl)
                currentYearWeek = yearWeek
        allWeeklyJpxDataList.append(JpxBrandData.JpxBrandData(adjdl.brandData,weeklyJpxDataList))
    return allWeeklyJpxDataList

def calcAllMonthlyJpxData(allDailyJpxDataList):
    allMonthlyJpxDataList = list()
    for i,adjdl in enumerate(allDailyJpxDataList):
        currentYearMonth = '-1'
        monthlyJpxDataList = list()
        monthlyJpxData = None
        prev = None
        for j,jdl in enumerate(adjdl.getJpxDataList()):
            yearMonth = CommonPackage.getYearMonth(jdl.date)
            if currentYearMonth == yearMonth:
                monthlyJpxData.append(jdl)
            if currentYearMonth != yearMonth:
                if monthlyJpxData != None:
                    monthlyJpxData.reCalc(prev)
                    prev = monthlyJpxData
                    monthlyJpxDataList.append(monthlyJpxData)
                monthlyJpxData = MonthlyJpxData.MonthlyJpxData(yearMonth,jdl)
                currentYearMonth = yearMonth
        allMonthlyJpxDataList.append(JpxBrandData.JpxBrandData(adjdl.brandData,monthlyJpxDataList))
    return allMonthlyJpxDataList

def writeDailyXlsx(allDailyJpxDataList):
    xlsxPath = os.path.join(CommonPackage.dataDir,datetime.datetime.today().strftime('%Y-%m-%d')+"-daily.xlsx")

    workbook = openpyxl.Workbook()
    worksheet = workbook['Sheet']
    workbook.remove(worksheet)
    print(xlsxPath)

    for i,adjdl in enumerate(allDailyJpxDataList):
        worksheet = workbook.create_sheet(title=adjdl.getBrandCode())
        DailyJpxData.DailyJpxData.writeJpxHeader(worksheet,1)
        for j,jdl in enumerate(adjdl.getJpxDataList(),2):
            jdl.writeJpxData(worksheet,j)
    workbook.save(xlsxPath)
    workbook.close()          

def writeWeeklyXlsx(allWeeklyJpxDataList):
    xlsxPath = os.path.join(CommonPackage.dataDir,datetime.datetime.today().strftime('%Y-%m-%d')+"-weekly.xlsx")

    workbook = openpyxl.Workbook()
    worksheet = workbook['Sheet']
    workbook.remove(worksheet)
    print(xlsxPath)

    for i,awjdl in enumerate(allWeeklyJpxDataList):
        worksheet = workbook.create_sheet(title=awjdl.getBrandCode())
        WeeklyJpxData.WeeklyJpxData.writeJpxHeader(worksheet,1)
        for j,jdl in enumerate(awjdl.getJpxDataList(),2):
            jdl.writeJpxData(worksheet,j)
    workbook.save(xlsxPath)
    workbook.close()          
        
def writeMonthlyXlsx(allMonthlyJpxDataList):
    xlsxPath = os.path.join(CommonPackage.dataDir,datetime.datetime.today().strftime('%Y-%m-%d')+"-monthly.xlsx")

    workbook = openpyxl.Workbook()
    worksheet = workbook['Sheet']
    workbook.remove(worksheet)
    print(xlsxPath)

    for i,awjdl in enumerate(allMonthlyJpxDataList):
        worksheet = workbook.create_sheet(title=awjdl.getBrandCode())
        MonthlyJpxData.MonthlyJpxData.writeJpxHeader(worksheet,1)
        for j,jdl in enumerate(awjdl.getJpxDataList(),2):
            jdl.writeJpxData(worksheet,j)
    workbook.save(xlsxPath)
    workbook.close()          

if __name__ == '__main__':

    brandData = [Brand.BrandData(info) for info in Brand.BrandData.getBrandInfo(idToken)]
    tradingCalender = TradingCalender.TradingCalender(idToken)
    print(f'Phase1 Start')
    start = datetime.datetime.now()
    length = len(brandData)
    for iter in range(math.ceil(length/CommonPackage.numOfThreads)):
        process = list()
        nextIter = CommonPackage.getNextInterIter(length,iter,CommonPackage.numOfThreads)
        for thread in range(nextIter):
            process.append(Process(target=dailyCheck,args=(idToken,brandData[iter*CommonPackage.numOfThreads+thread],1)))
        for p in process:
            p.start()
        for p in process:
            p.join()
    end = datetime.datetime.now()
    print(f'Phase1 End ElapsedTime = {end-start}')
    start = datetime.datetime.now()

    allDailyJpxDataList = readDailyAllJpxData(brandData)
    writeDailyXlsx(allDailyJpxDataList)

    end = datetime.datetime.now()
    print(f'Phase2 End ElapsedTime = {end-start}')
    start = datetime.datetime.now()

    allWeeklyJpxDataList = calcAllWeeklyJpxData(allDailyJpxDataList)
    writeWeeklyXlsx(allWeeklyJpxDataList)

    end = datetime.datetime.now()
    print(f'Phase3 End ElapsedTime = {end-start}')
    start = datetime.datetime.now()

    allMonthlyJpxDataList = calcAllMonthlyJpxData(allDailyJpxDataList)
    writeMonthlyXlsx(allMonthlyJpxDataList)

    end = datetime.datetime.now()
    print(f'Phase4 End ElapsedTime = {end-start}')
    start = datetime.datetime.now()
